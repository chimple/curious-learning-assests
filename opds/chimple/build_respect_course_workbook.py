#!/usr/bin/env python3
"""Build the RESPECT lesson workbook used by the Chimple/Lido script.

Three source modes are supported:

1) Dashboard API mode (recommended for fresh data)
   The script takes a curriculum ID plus grade IDs and calls a dashboard/API
   endpoint that returns course-detail rows. Configure the endpoint using
   --api-url or CHIMPLE_DASHBOARD_API_URL.

2) Dashboard XLSX mode
   The script can rebuild the five output sheets from an existing Chimple
   dashboard export. If an ``All Courses`` sheet is present, it uses it to
   validate/repair grade/course membership instead of blindly copying tabs.

3) Direct Supabase mode
   The script reads the public course, chapter, chapter_lesson, and lesson
   tables directly. Supply the URL and key as arguments or environment
   variables; credentials are never stored in this script.

The output workbook always contains exactly these sheets:
  English Grade 1, English Grade 2, Maths Grade 1, Maths Grade 2, Digital Skills
and exactly these columns:
  lesson_id, title, lesson_name, cocos_lesson_code, lido_lesson_id,
  cocosChapterCode, Asset Link
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_NAME = "Respect Course Latest All Course Details From dashboard.xlsx"
SHEETS = [
    "English Grade 1",
    "English Grade 2",
    "Maths Grade 1",
    "Maths Grade 2",
    "Digital Skills",
]
COLUMNS = [
    "lesson_id",
    "title",
    "lesson_name",
    "cocos_lesson_code",
    "lido_lesson_id",
    "cocosChapterCode",
    "Asset Link",
]

# Used only to identify the intended course inside a broad dashboard XLSX export.
SHEET_HINTS = {
    "English Grade 1": ("Grade 1", "English"),
    "English Grade 2": ("Grade 2", "English"),
    "Maths Grade 1": ("Grade 1", "Maths"),
    "Maths Grade 2": ("Grade 2", "Maths"),
    "Digital Skills": ("Below Grade 1", "Digital Skills"),
}

ASSET_BASE = "https://raw.githubusercontent.com/chimple/chimple-zips/main"

ALIASES = {
    "lesson_id": ["lesson_id", "lessonId", "lesson_doc_id", "doc_id"],
    "title": ["title", "lesson_name", "lessonName", "name"],
    "lesson_name": ["lesson_name", "lessonName", "title", "name"],
    "cocos_lesson_code": [
        "cocos_lesson_code",
        "cocosLessonCode",
        "cocos_lesson_id",
        "cocosLessonId",
        "id",
    ],
    "lido_lesson_id": ["lido_lesson_id", "lidoLessonId"],
    "cocosChapterCode": [
        "cocosChapterCode",
        "cocos_chapter_code",
        "cocos_chapter_id",
        "cocosChapterId",
    ],
    "Asset Link": ["Asset Link", "asset_link", "assetLink", "zip_url", "zipUrl"],
    "course_id": ["course_id", "courseId"],
    "course_name": ["course_name", "courseName", "subject_name", "subjectName"],
    "course_grade_name": ["course_grade_name", "grade_name", "gradeName"],
}


def clean_key(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def pick(row: Mapping[str, Any], canonical: str) -> Any:
    for key in ALIASES.get(canonical, [canonical]):
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def derive_chapter_code(cocos_code: str) -> str:
    """en0400 -> en04, maths1604 -> maths16, puzzle0001 -> puzzle00."""
    code = cocos_code.strip()
    m = re.match(r"^(.*?)(\d{4})$", code)
    if not m:
        return ""
    return f"{m.group(1)}{m.group(2)[:2]}"


def make_asset_link(cocos_code: str) -> str:
    return f"{ASSET_BASE}/{cocos_code}.zip"


def normalize_lesson(row: Mapping[str, Any], generate_asset: bool = True) -> Optional[Dict[str, str]]:
    lesson_id = clean_text(pick(row, "lesson_id"))
    if not lesson_id:
        # Chapter/group heading: explicitly excluded by requirement.
        return None

    cocos_code = clean_text(pick(row, "cocos_lesson_code"))
    lido_lesson_id = clean_text(pick(row, "lido_lesson_id"))
    if not cocos_code and not lido_lesson_id:
        return None

    name = clean_text(pick(row, "lesson_name")) or clean_text(pick(row, "title"))
    if not name:
        return None

    asset = clean_text(pick(row, "Asset Link"))
    # Cocos bundles are downloadable ZIPs. A Lido-only lesson has no Cocos ZIP
    # and must remain publishable so Cuba can use its Lido compatibility path.
    if cocos_code and not asset and generate_asset:
        asset = make_asset_link(cocos_code)
    if cocos_code and not asset:
        return None

    chapter_code = clean_text(pick(row, "cocosChapterCode"))
    if cocos_code and not chapter_code:
        chapter_code = derive_chapter_code(cocos_code)

    return {
        "lesson_id": lesson_id,
        "title": name,
        "lesson_name": name,
        "cocos_lesson_code": cocos_code,
        "lido_lesson_id": lido_lesson_id,
        "cocosChapterCode": chapter_code,
        "Asset Link": asset,
    }


def verify_asset(url: str, timeout: float = 12.0) -> bool:
    """Return True only when the public ZIP URL responds successfully."""
    try:
        # Some hosts do not support HEAD reliably, so fall back to a tiny GET.
        r = requests.head(url, allow_redirects=True, timeout=timeout)
        if r.status_code < 400:
            return True
        r = requests.get(
            url,
            allow_redirects=True,
            timeout=timeout,
            headers={"Range": "bytes=0-0"},
            stream=True,
        )
        return r.status_code in (200, 206)
    except requests.RequestException:
        return False


def dedupe_lessons(rows: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    result: List[Dict[str, str]] = []
    seen: set[str] = set()
    for row in rows:
        lid = row["lesson_id"]
        if lid in seen:
            continue
        seen.add(lid)
        result.append(row)
    return result


def parse_grade_args(items: Sequence[str]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for item in items:
        if "=" not in item:
            raise ValueError(f"Invalid --grade-id {item!r}; use 'SHEET NAME=GRADE_ID'.")
        sheet, grade_id = item.split("=", 1)
        sheet, grade_id = sheet.strip(), grade_id.strip()
        if sheet not in SHEETS:
            raise ValueError(f"Unknown sheet in --grade-id: {sheet!r}")
        if not grade_id:
            raise ValueError(f"Empty grade ID for {sheet!r}")
        mapping[sheet] = grade_id
    return mapping


def extract_json_rows(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if isinstance(payload, dict):
        for key in ("rows", "data", "results", "items", "course_details"):
            value = payload.get(key)
            if isinstance(value, list):
                return [x for x in value if isinstance(x, dict)]
        # Some APIs return {"data": {"rows": [...]}}
        data = payload.get("data")
        if isinstance(data, dict):
            return extract_json_rows(data)
    raise ValueError("API response does not contain a recognizable row list.")


def subject_matches(sheet: str, row: Mapping[str, Any]) -> bool:
    """Keep the row in the subject sheet requested by the caller.

    If the API supplies course/subject name, use it. Otherwise fall back to
    cocos code prefixes for English/Maths. Digital Skills can contain several
    game prefixes (puzzle, pictureboard, drawshape, matchingcard, ...), so the
    API should ideally provide course_name/subject_name for that sheet.
    """
    expected_subject = SHEET_HINTS[sheet][1].lower()
    course_name = clean_text(pick(row, "course_name")).lower()
    if course_name:
        if expected_subject == "maths":
            return course_name in ("maths", "math", "mathematics") or "math" in course_name
        return expected_subject in course_name

    code = clean_text(pick(row, "cocos_lesson_code")).lower()
    if sheet.startswith("English"):
        return code.startswith("en")
    if sheet.startswith("Maths"):
        return code.startswith("maths")
    return True


def fetch_api_rows(
    api_url: str,
    curriculum_id: str,
    grade_id: str,
    token: Optional[str],
    timeout: float,
) -> List[Dict[str, Any]]:
    headers = {"Accept": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    response = requests.get(
        api_url,
        params={"curriculum_id": curriculum_id, "grade_id": grade_id},
        headers=headers,
        timeout=timeout,
    )
    response.raise_for_status()
    return extract_json_rows(response.json())


def build_from_api(args: argparse.Namespace) -> Dict[str, List[Dict[str, str]]]:
    grade_map = grade_map_from_args(args, "API")
    if not args.curriculum_id:
        raise ValueError("API mode requires --curriculum-id.")

    api_url = args.api_url or os.getenv("CHIMPLE_DASHBOARD_API_URL", "")
    if not api_url:
        raise ValueError(
            "API mode needs --api-url or CHIMPLE_DASHBOARD_API_URL. "
            "Point it at the Chimple dashboard/course-details endpoint that accepts "
            "curriculum_id and grade_id query parameters."
        )
    token = args.api_token or os.getenv("CHIMPLE_DASHBOARD_API_TOKEN")

    cache: Dict[str, List[Dict[str, Any]]] = {}
    out: Dict[str, List[Dict[str, str]]] = {}
    for sheet in SHEETS:
        gid = grade_map[sheet]
        if gid not in cache:
            cache[gid] = fetch_api_rows(
                api_url=api_url,
                curriculum_id=args.curriculum_id,
                grade_id=gid,
                token=token,
                timeout=args.timeout,
            )
        normalized: List[Dict[str, str]] = []
        for raw in cache[gid]:
            if not subject_matches(sheet, raw):
                continue
            row = normalize_lesson(raw, generate_asset=True)
            if row:
                normalized.append(row)
        out[sheet] = dedupe_lessons(normalized)
    return out


def grade_map_from_args(args: argparse.Namespace, source_name: str) -> Dict[str, str]:
    """Resolve shared grade shortcuts for the two live-data source modes."""
    grade_map = parse_grade_args(args.grade_id or [])
    if args.grade_1_id:
        grade_map.setdefault("English Grade 1", args.grade_1_id)
        grade_map.setdefault("Maths Grade 1", args.grade_1_id)
    if args.grade_2_id:
        grade_map.setdefault("English Grade 2", args.grade_2_id)
        grade_map.setdefault("Maths Grade 2", args.grade_2_id)
    if args.digital_grade_id:
        grade_map.setdefault("Digital Skills", args.digital_grade_id)
    missing = [s for s in SHEETS if s not in grade_map]
    if missing:
        raise ValueError(f"{source_name} mode requires --grade-id for: " + ", ".join(missing))
    return grade_map


def chunks(items: Sequence[str], size: int = 400) -> Iterable[Sequence[str]]:
    for index in range(0, len(items), size):
        yield items[index : index + size]


def supabase_rows(
    base_url: str,
    table: str,
    params: Mapping[str, str],
    key: str,
    timeout: float,
) -> List[Dict[str, Any]]:
    """Fetch every page from a Supabase REST table without exposing the key."""
    endpoint = f"{base_url.rstrip('/')}/rest/v1/{table}"
    headers = {
        "Accept": "application/json",
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Range-Unit": "items",
    }
    rows: List[Dict[str, Any]] = []
    start, page_size = 0, 1000
    while True:
        page_headers = {**headers, "Range": f"{start}-{start + page_size - 1}"}
        response = requests.get(endpoint, params=params, headers=page_headers, timeout=timeout)
        response.raise_for_status()
        page = extract_json_rows(response.json())
        rows.extend(page)
        if len(page) < page_size:
            return rows
        start += page_size


def fetch_supabase_course_rows(
    base_url: str,
    key: str,
    curriculum_id: str,
    grade_id: str,
    timeout: float,
) -> List[Dict[str, Any]]:
    courses = supabase_rows(
        base_url,
        "course",
        {
            "select": "id,name,code,sort_index",
            "curriculum_id": f"eq.{curriculum_id}",
            "grade_id": f"eq.{grade_id}",
            "is_deleted": "eq.false",
            "order": "sort_index.asc",
        },
        key,
        timeout,
    )
    course_by_id = {clean_text(course.get("id")): course for course in courses}
    if not course_by_id:
        return []

    chapters: List[Dict[str, Any]] = []
    for course_ids in chunks(list(course_by_id)):
        chapters.extend(
            supabase_rows(
                base_url,
                "chapter",
                {
                    "select": "id,course_id,name,sort_index",
                    "course_id": f"in.({','.join(course_ids)})",
                    "is_deleted": "eq.false",
                    "order": "sort_index.asc",
                },
                key,
                timeout,
            )
        )
    chapter_by_id = {clean_text(chapter.get("id")): chapter for chapter in chapters}
    if not chapter_by_id:
        return []

    # Query the join table, rather than lesson directly, because lesson membership
    # and ordering are defined by chapter_lesson in Cuba as well.
    course_rows: List[Dict[str, Any]] = []
    for chapter_ids in chunks(list(chapter_by_id)):
        links = supabase_rows(
            base_url,
            "chapter_lesson",
            {
                "select": (
                    "chapter_id,lesson_id,sort_index,"
                    "lesson:lesson_id(id,name,cocos_lesson_id,lido_lesson_id,"
                    "cocos_chapter_code,is_deleted)"
                ),
                "chapter_id": f"in.({','.join(chapter_ids)})",
                "is_deleted": "eq.false",
                "order": "sort_index.asc",
            },
            key,
            timeout,
        )
        for link in links:
            lesson = link.get("lesson")
            chapter = chapter_by_id.get(clean_text(link.get("chapter_id")))
            if not isinstance(lesson, dict) or not chapter or lesson.get("is_deleted"):
                continue
            course = course_by_id.get(clean_text(chapter.get("course_id")))
            if not course:
                continue
            course_rows.append(
                {
                    "course_id": course.get("id"),
                    "course_name": course.get("name") or course.get("code"),
                    "lesson_id": lesson.get("id") or link.get("lesson_id"),
                    "lesson_name": lesson.get("name"),
                    "cocos_lesson_code": lesson.get("cocos_lesson_id"),
                    "lido_lesson_id": lesson.get("lido_lesson_id"),
                    "cocosChapterCode": lesson.get("cocos_chapter_code"),
                    "chapter_index": chapter.get("sort_index"),
                    "lesson_index": link.get("sort_index"),
                }
            )
    return course_rows


def build_from_supabase(args: argparse.Namespace) -> Dict[str, List[Dict[str, str]]]:
    grade_map = grade_map_from_args(args, "Supabase")
    if not args.curriculum_id:
        raise ValueError("Supabase mode requires --curriculum-id.")

    base_url = args.supabase_url or os.getenv("SUPABASE_URL", "")
    key = args.supabase_key or os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_ANON_KEY", "")
    if not base_url or not key:
        raise ValueError(
            "Supabase mode needs --supabase-url / SUPABASE_URL and "
            "--supabase-key / SUPABASE_SERVICE_ROLE_KEY."
        )

    cache: Dict[str, List[Dict[str, Any]]] = {}
    out: Dict[str, List[Dict[str, str]]] = {}
    for sheet in SHEETS:
        gid = grade_map[sheet]
        if gid not in cache:
            cache[gid] = fetch_supabase_course_rows(
                base_url=base_url,
                key=key,
                curriculum_id=args.curriculum_id,
                grade_id=gid,
                timeout=args.timeout,
            )
        normalized: List[Dict[str, str]] = []
        for raw in cache[gid]:
            if not subject_matches(sheet, raw):
                continue
            row = normalize_lesson(raw, generate_asset=True)
            if row:
                normalized.append(row)
        out[sheet] = dedupe_lessons(normalized)
    return out


def worksheet_dict_rows(ws) -> List[Dict[str, Any]]:
    headers = [clean_key(c.value) for c in ws[1]]
    rows: List[Dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        d: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            d[header] = values[idx] if idx < len(values) else None
        rows.append(d)
    return rows


def choose_course_from_all_courses(
    sheet: str,
    all_rows: List[Dict[str, Any]],
    current_rows: List[Dict[str, Any]],
) -> Optional[str]:
    expected_grade, expected_subject = SHEET_HINTS[sheet]
    candidates: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in all_rows:
        cid = clean_text(pick(row, "course_id"))
        grade = clean_text(pick(row, "course_grade_name"))
        course = clean_text(pick(row, "course_name"))
        if cid and grade == expected_grade and course == expected_subject:
            candidates[cid].append(row)
    if not candidates:
        return None

    current_ids = {
        clean_text(pick(r, "lesson_id")) for r in current_rows if clean_text(pick(r, "lesson_id"))
    }
    current_codes = {
        clean_text(pick(r, "cocos_lesson_code"))
        for r in current_rows
        if clean_text(pick(r, "cocos_lesson_code"))
    }

    scored: List[Tuple[int, int, int, str]] = []
    for cid, rows in candidates.items():
        ids = {clean_text(pick(r, "lesson_id")) for r in rows if clean_text(pick(r, "lesson_id"))}
        codes = {
            clean_text(pick(r, "cocos_lesson_code"))
            for r in rows
            if clean_text(pick(r, "cocos_lesson_code"))
        }
        # Prefer exact internal-ID overlap. If a tab was mistakenly copied from
        # another grade, code overlap repairs it while retaining the right course.
        scored.append((len(ids & current_ids), len(codes & current_codes), len(rows), cid))
    scored.sort(reverse=True)
    return scored[0][3]


def build_from_xlsx(path: Path) -> Tuple[Dict[str, List[Dict[str, str]]], Dict[str, str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise ValueError("Source workbook is missing required sheets: " + ", ".join(missing))

    all_rows = worksheet_dict_rows(wb["All Courses"]) if "All Courses" in wb.sheetnames else []
    output: Dict[str, List[Dict[str, str]]] = {}
    selected_courses: Dict[str, str] = {}

    for sheet in SHEETS:
        current_rows = worksheet_dict_rows(wb[sheet])
        source_rows = current_rows

        if all_rows:
            course_id = choose_course_from_all_courses(sheet, all_rows, current_rows)
            if course_id:
                selected_courses[sheet] = course_id
                source_rows = [r for r in all_rows if clean_text(pick(r, "course_id")) == course_id]

        # Enrich dashboard-wide rows from the curated sheet when the broad
        # export is missing a human-readable title or an already-known asset URL.
        current_by_id = {
            clean_text(pick(r, "lesson_id")): r
            for r in current_rows
            if clean_text(pick(r, "lesson_id"))
        }
        current_by_code = {
            clean_text(pick(r, "cocos_lesson_code")): r
            for r in current_rows
            if clean_text(pick(r, "cocos_lesson_code"))
        }

        normalized = []
        for raw in source_rows:
            enriched = dict(raw)
            fallback = (
                current_by_id.get(clean_text(pick(raw, "lesson_id")))
                or current_by_code.get(clean_text(pick(raw, "cocos_lesson_code")))
            )
            if fallback:
                for canonical in ("title", "lesson_name", "cocosChapterCode", "Asset Link"):
                    if not clean_text(pick(enriched, canonical)):
                        value = pick(fallback, canonical)
                        if value not in (None, ""):
                            enriched[canonical] = value
            row = normalize_lesson(enriched, generate_asset=True)
            if row:
                normalized.append(row)
        output[sheet] = dedupe_lessons(normalized)

    return output, selected_courses


def write_workbook(data: Mapping[str, Sequence[Mapping[str, str]]], output: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in SHEETS:
        ws = wb.create_sheet(sheet)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
        ws.sheet_view.showGridLines = False

        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, item in enumerate(data.get(sheet, []), start=2):
            for col_idx, header in enumerate(COLUMNS, start=1):
                cell = ws.cell(row_idx, col_idx, item.get(header, ""))
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if header == "Asset Link" and cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

        widths = {
            "A": 24,
            "B": 34,
            "C": 34,
            "D": 24,
            "E": 24,
            "F": 78,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.row_dimensions[1].height = 24
        ws.auto_filter.ref = f"A1:F{max(1, ws.max_row)}"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def validate_output(data: Mapping[str, Sequence[Mapping[str, str]]], verify_assets: bool) -> None:
    errors: List[str] = []
    for sheet in SHEETS:
        ids = set()
        for i, row in enumerate(data.get(sheet, []), start=2):
            for col in ("lesson_id", "title", "lesson_name"):
                if not clean_text(row.get(col)):
                    errors.append(f"{sheet}!{i}: empty {col}")
            cocos_code = clean_text(row.get("cocos_lesson_code"))
            lido_lesson_id = clean_text(row.get("lido_lesson_id"))
            if not cocos_code and not lido_lesson_id:
                errors.append(f"{sheet}!{i}: missing Cocos and Lido lesson IDs")
            if cocos_code:
                # Chapter code is optional catalogue metadata. The bundle URL is
                # the required Cocos playback dependency.
                if not clean_text(row.get("Asset Link")):
                    errors.append(f"{sheet}!{i}: empty Asset Link for Cocos lesson")
            lid = row.get("lesson_id", "")
            if lid in ids:
                errors.append(f"{sheet}!{i}: duplicate lesson_id {lid}")
            ids.add(lid)
            if verify_assets and row.get("Asset Link") and not verify_asset(row["Asset Link"]):
                errors.append(f"{sheet}!{i}: asset unavailable: {row['Asset Link']}")
    if errors:
        preview = "\n".join(errors[:30])
        more = "" if len(errors) <= 30 else f"\n... and {len(errors)-30} more"
        raise ValueError(f"Output validation failed:\n{preview}{more}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--source-xlsx", type=Path, help="Existing Chimple dashboard XLSX export.")
    source.add_argument("--api", action="store_true", help="Fetch fresh rows using curriculum/grade IDs.")
    source.add_argument(
        "--supabase",
        action="store_true",
        help="Fetch fresh course, chapter, and lesson data directly from Supabase REST.",
    )

    parser.add_argument("--curriculum-id", help="Chimple curriculum ID (API or Supabase mode).")
    parser.add_argument(
        "--grade-id",
        action="append",
        default=[],
        metavar="SHEET=GRADE_ID",
        help=(
            "Per-sheet grade ID override. Repeat as needed. Example: "
            "--grade-id 'English Grade 1=abc123'"
        ),
    )
    parser.add_argument(
        "--grade-1-id",
        help="Shortcut: grade ID used for both English Grade 1 and Maths Grade 1.",
    )
    parser.add_argument(
        "--grade-2-id",
        help="Shortcut: grade ID used for both English Grade 2 and Maths Grade 2.",
    )
    parser.add_argument(
        "--digital-grade-id",
        help="Grade ID used for Digital Skills (for example the Below Grade 1 ID).",
    )
    parser.add_argument("--api-url", help="Course-details API URL; or set CHIMPLE_DASHBOARD_API_URL.")
    parser.add_argument("--api-token", help="Bearer token; or set CHIMPLE_DASHBOARD_API_TOKEN.")
    parser.add_argument("--supabase-url", help="Supabase project URL; or set SUPABASE_URL.")
    parser.add_argument(
        "--supabase-key",
        help="Supabase service-role/anon key; or set SUPABASE_SERVICE_ROLE_KEY/SUPABASE_ANON_KEY.",
    )
    parser.add_argument("--timeout", type=float, default=30.0, help="HTTP timeout in seconds.")
    parser.add_argument(
        "--verify-assets",
        action="store_true",
        help="Check every public ZIP URL and fail if any asset cannot be reached.",
    )
    parser.add_argument("--output", type=Path, default=Path(OUTPUT_NAME))
    args = parser.parse_args()

    if args.source_xlsx:
        data, selected = build_from_xlsx(args.source_xlsx)
        if selected:
            print("Selected dashboard course IDs:")
            for s in SHEETS:
                print(f"  {s}: {selected.get(s, '(not resolved)')}")
    elif args.supabase:
        data = build_from_supabase(args)
    else:
        data = build_from_api(args)

    validate_output(data, verify_assets=args.verify_assets)
    write_workbook(data, args.output)

    print(f"Wrote: {args.output}")
    print("Lesson rows:")
    for sheet in SHEETS:
        print(f"  {sheet}: {len(data.get(sheet, []))}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
