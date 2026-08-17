import os
import json
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from openpyxl import load_workbook
from urllib.parse import urlencode, urljoin

# Configuration
EXCEL_FILE = 'Respect Course Latest All Course Details From dashboard.xlsx'
# EXCEL_FILE = 'sheet (1).xlsx'

BASE_URL = 'https://chimple-respect.web.app/'  # Base URL for your OPDS catalog
TYPE_OPDS = 'application/opds+json'
PUB_TYPE = 'application/opds-publication+json'
DEFAULT_COLLECTION_FILENAME = 'opds.json'
# RESPECT must launch the installed Android app without relying on website App Link verification.
CHIMPLE_LESSON_LAUNCH_BASE = 'chimple://respect/launch'
SKIP_SHEETS = {'All Courses', 'Sheet4', 'Sheet5'}  # adjust as needed
GRADE_KEYS = {
    'English Grade 1': 'en_g1',
    'English Grade 2': 'en_g2',
    'Maths Grade 1': 'maths_g1',
    'Maths Grade 2': 'maths_g2',
    'Digital Skills': 'puzzle',
}
GRADE_ICON_URLS = {
    'english': 'https://pub-3a82c17429da40d1989930ae7eb2f2d1.r2.dev/course-icons/6796ee8f-a237-42a3-beda-94d61a7139a1-4de8f55c.webp',
    'maths': 'https://pub-3a82c17429da40d1989930ae7eb2f2d1.r2.dev/course-icons/e5cda413-2f6c-485d-8985-27e3b6aad0d5-6cca9deb.webp',
    'digital': 'https://pub-3a82c17429da40d1989930ae7eb2f2d1.r2.dev/course-icons/19bb079f-bc69-44e4-bc1d-0b77f2683b6c-5a17f2ef.webp',
}

# Output folders
OUTPUT_DIR = 'public'
GRADE_DIR = os.path.join(OUTPUT_DIR, 'grades')
LESSON_DIR = os.path.join(OUTPUT_DIR, 'lessons')
ICONS_DIR = os.path.join(OUTPUT_DIR, 'images', 'icons')

def get_lesson_launch_url(activity_id, chimple_lesson_id=''):
    """Keep the xAPI activity canonical while supplying Cuba's playable bundle ID."""
    launch_parameters = {'activity_id': activity_id}
    if chimple_lesson_id:
        launch_parameters['chimple_lesson_id'] = chimple_lesson_id
    return f'{CHIMPLE_LESSON_LAUNCH_BASE}?{urlencode(launch_parameters)}'

def clean_lesson_value(value):
    value = str(value or '').strip()
    return '' if value.lower() in {'nan', 'none'} else value


def get_lido_lesson_id(cocos_lesson_code, lido_lesson_id=''):
    """Return the Lido bundle ID used by Cuba and the browser player."""
    lido_lesson_id = clean_lesson_value(lido_lesson_id)
    if lido_lesson_id.startswith('LIDO_'):
        return lido_lesson_id

    cocos_lesson_code = clean_lesson_value(cocos_lesson_code)
    if not cocos_lesson_code or cocos_lesson_code.lower() == 'default':
        return None

    return (
        f'LIDO_{cocos_lesson_code}_en'
        if cocos_lesson_code.lower().startswith('maths')
        else f'LIDO_{cocos_lesson_code}'
    )


def get_lido_browser_launch_url(cocos_lesson_code, lido_lesson_id=''):
    lido_lesson_id = get_lido_lesson_id(cocos_lesson_code, lido_lesson_id)
    return f'https://chimple.cc/{lido_lesson_id}' if lido_lesson_id else None

os.makedirs(GRADE_DIR, exist_ok=True)
os.makedirs(LESSON_DIR, exist_ok=True)

def get_image_path(cocos_lesson_code):
    """Get the image path for a lesson, fallback to default if not found"""
    if not cocos_lesson_code or cocos_lesson_code == 'default':
        return 'default.png'
    
    # Try to find the image with the cocos_lesson_code
    possible_extensions = ['.png', '.png', '.png', '.webp']
    for ext in possible_extensions:
        image_path = os.path.join(ICONS_DIR, f"{cocos_lesson_code}{ext}")
        if os.path.exists(image_path):
            return f"{cocos_lesson_code}{ext}"
    
    # If not found, return default
    return 'default.png'

# Load workbook
print(f"Loading workbook: {EXCEL_FILE}")
wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)

# --- Gather navigation and learning units ---
navigation = []
learning_units = []
for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue
    filename = GRADE_KEYS.get(sheet_name, sheet_name.replace(' ', '').lower()) + '.json'
    
    # Course icons are supplied by the curriculum owner and shared by each subject's grades.
    if 'english' in sheet_name.lower():
        icon = GRADE_ICON_URLS['english']
    elif 'maths' in sheet_name.lower():
        icon = GRADE_ICON_URLS['maths']
    elif 'digital' in sheet_name.lower():
        icon = GRADE_ICON_URLS['digital']
    else:
        icon = BASE_URL + 'images/icons/default.png'
    
    nav_obj = {
        'href': BASE_URL + 'grades/' + filename,
        'title': sheet_name,
        'type': TYPE_OPDS,
        'alternate': [
            {
                'href': icon,
                'rel': "icon",
                'type': "image/webp" if icon.endswith('.webp') else "image/png",
                'title': sheet_name
            }
        ]
    }
    navigation.append(nav_obj)
    learning_units.append({
        "id": filename.replace('.json', ''),
        "title": sheet_name,
        "href": nav_obj['href'],
        "type": TYPE_OPDS
    })

# --- Generate opds.json (OPDS catalog) ---
opds_catalog = {
    "metadata": {"title": "Chimple Learning"},
    "links": [{"rel": "self", "href": urljoin(BASE_URL, 'opds.json'), "type": TYPE_OPDS}],
    "navigation": navigation
}
with open(os.path.join(OUTPUT_DIR, 'opds.json'), 'w', encoding='utf-8') as f:
    json.dump(opds_catalog, f, indent=2)
print(f"Generated opds.json with {len(navigation)} grades.")

# --- Generate index.json (RESPECT App Manifest) ---
respect_manifest = {
    "name": {"en-US": "Chimple Learning"},
    "description": {"en-US": "A collection of interactive learning units for children."},
    "license": "MIT",
    "website": BASE_URL,
    "icon": BASE_URL + "icon.webp",
    "learningUnits": BASE_URL + DEFAULT_COLLECTION_FILENAME,
    "defaultLaunchUri": navigation[0]['href'] if navigation else BASE_URL,
    "android": {
        "packageId": "org.chimple.cuba",
        "stores": ["https://play.google.com/store/apps/details?id=org.chimple.cuba"]
    },
    "web": {
        "url": BASE_URL
    }
}
with open(os.path.join(OUTPUT_DIR, 'index.json'), 'w', encoding='utf-8') as f:
    json.dump(respect_manifest, f, indent=2)
print(f"Generated index.json as RESPECT App Manifest.")

launchable_app_manifest = {
    "metadata": {
        "@type": "https://id.openeel.org/schema/launchable-app",
        "title": "Chimple Kids Learning",
        "description": "Interactive learning units from Chimple.",
        "author": {"name": "Chimple Learning"},
        "identifier": f"{BASE_URL}app",
        "language": "en",
        "modified": datetime.now(timezone.utc).isoformat(),
    },
    "links": [
        {
            "rel": "self",
            "href": f"{BASE_URL}launchable-app.json",
            "type": PUB_TYPE,
        },
        {
            "rel": "collection",
            "href": f"{BASE_URL}{DEFAULT_COLLECTION_FILENAME}",
            "type": TYPE_OPDS,
        },
        {
            "rel": "https://id.openeel.org/rel/app-launch-uri",
            "href": "https://chimple.cc/",
        },
        {
            "rel": "https://id.openeel.org/rel/appstore-android",
            "href": "https://play.google.com/store/apps/details?id=org.chimple.bahama",
            "title": "Get it on Google Play",
        },
        {"rel": "terms-of-service", "href": "https://www.chimple.org/privacy-policy"},
        {"rel": "license", "href": "https://www.gnu.org/licenses/agpl-3.0.html"},
    ],
    "images": [
        {
            "href": "https://raw.githubusercontent.com/chimple/cuba/RESPECTify/public/assets/icons/favicon.png",
            "type": "image/png",
        }
    ],
}
with open(os.path.join(OUTPUT_DIR, 'launchable-app.json'), 'w', encoding='utf-8') as f:
    json.dump(launchable_app_manifest, f, indent=2)
print("Generated launchable-app.json.")

def create_lesson_manifest(lesson_data, lesson_id, title, asset_link):
    """Create a lesson manifest in Readium Web Publication Manifest format"""
    # Fix date format to RFC 3339 (UTC timezone)
    current_time = datetime.now(timezone.utc).isoformat()
    
    # Get image path
    cocos_lesson_code = clean_lesson_value(lesson_data.get('cocos_lesson_code'))
    lido_lesson_id = get_lido_lesson_id(
        cocos_lesson_code,
        lesson_data.get('lido_lesson_id'),
    )
    image_code = cocos_lesson_code or 'default'
    image_filename = get_image_path(image_code)
    browser_launch_url = get_lido_browser_launch_url(cocos_lesson_code, lido_lesson_id)
    print(f"Using image for lesson {lesson_id}: {image_filename}")

    
    # Create Readium Web Publication Manifest
    lesson_manifest = {
        "@context": [
            "https://readium.org/webpub-manifest/context.jsonld",
            "https://schema.org"
        ],
        "metadata": {
            "@type": "http://schema.org/Game",
            "title": title,
            "author": "Chimple",
            "identifier": f"{BASE_URL}activities/{lesson_id}",
            "language": "en",
            "modified": current_time,
            "published": current_time,
            "description": f"Interactive learning lesson: {title}",
            "subject": ["Education", "Learning"],
            "readingProgression": "ltr"
        },
        "links": [
            {
                "rel": "self",
                "href": f"{BASE_URL}lessons/{lesson_id}.json",
                "type": PUB_TYPE
            },
            {
                "rel": "https://id.openeel.org/rel/tincanxml",
                "href": f"{BASE_URL}lessons/{lesson_id}/tincan.xml",
                "type": "application/xml"
            },
            {
                "rel": "https://id.openeel.org/rel/launchable-app",
                "href": f"{BASE_URL}launchable-app.json",
                "type": PUB_TYPE
            }
        ],
        "images": [
            {
                "href": f"{BASE_URL}images/icons/{image_filename}",
                "type": "image/png",
                "height": 128,
                "width": 128
            }
        ],

        "readingOrder": [
            {
                "type": "text/html",
                "href": browser_launch_url or get_lesson_launch_url(f'{BASE_URL}activities/{lesson_id}'),
                "title": title
            }
        ],
        "resources": [
            {
                "type": "image/png",
                "href": f"{BASE_URL}images/icons/{image_filename}",
                "properties": {
                    "width": 128,
                    "height": 128
                }
            }
        ]
    }

    # Lido-only lessons have no Cocos ZIP. Do not publish an empty offline
    # resource; Cuba will resolve lido_lesson_id when no Cocos ID is present.
    if cocos_lesson_code and asset_link:
        lesson_manifest['resources'].append({
            "type": "application/zip",
            "href": asset_link,
            "properties": {
                "contains": ["application/xhtml+xml", "text/css", "image/*"]
            }
        })
    
    if browser_launch_url:
        lesson_manifest['links'].append({
            "rel": "http://opds-spec.org/acquisition/open-access",
            "href": browser_launch_url,
            "type": "text/html"
        })

    return lesson_manifest


def create_tincan_xml(lesson_id, title, chimple_lesson_id):
    """Create the Rustici launch metadata for a lesson's canonical xAPI activity."""
    lesson_dir = os.path.join(LESSON_DIR, lesson_id)
    os.makedirs(lesson_dir, exist_ok=True)

    namespace = 'http://projecttincan.com/tincan.xsd'
    ET.register_namespace('', namespace)
    root = ET.Element(f'{{{namespace}}}tincan')
    activities = ET.SubElement(root, 'activities')
    activity = ET.SubElement(
        activities,
        'activity',
        {
            'id': f'{BASE_URL}activities/{lesson_id}',
            'type': 'http://activitystrea.ms/schema/1.0/game',
        },
    )
    ET.SubElement(activity, 'name').text = title
    description = ET.SubElement(activity, 'description', {'lang': 'en-US'})
    description.text = f'Chimple learning activity: {title}'
    launch = ET.SubElement(activity, 'launch', {'lang': 'en-us'})
    launch.text = get_lesson_launch_url(
        f'{BASE_URL}activities/{lesson_id}',
        chimple_lesson_id,
    )

    ET.ElementTree(root).write(
        os.path.join(lesson_dir, 'tincan.xml'),
        encoding='utf-8',
        xml_declaration=True,
    )

# --- Process sheets ---
for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue
    
    print(f"\nProcessing sheet: {sheet_name}")
    ws = wb[sheet_name]
    # Keep the published grade feed URLs stable for existing RESPECT clients.
    grade_key = GRADE_KEYS.get(sheet_name, sheet_name.replace(' ', '').lower())
    grade_file = f"{grade_key}.json"
    publications = []
    
    headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Headers (columns) fetched from sheet '{sheet_name}': {headers}")
    for i, h in enumerate(headers):
        print(f"Header {i}: '{h}' (len={len(h)})")

    row_count = 0
    valid_lessons = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        data = dict(zip(headers, row))
        data = {k: (v.strip() if isinstance(v, str) else v) for k, v in data.items()}
        if row_count <= 5:
            print(f"Row {row_count} keys: {list(data.keys())}")
            print(f"Row {row_count} values: {list(data.values())}")
            print(f"Row {row_count} data: {data}")
        lesson_id_value = data.get('lesson_id')
        if lesson_id_value is None:
            print(f"Skipping row {row_count}: Missing lesson_id")
            continue
        lesson_id = str(lesson_id_value).strip()
        if not lesson_id or lesson_id.lower() == 'nan':
            print(f"Skipping row {row_count}: Missing lesson_id (value: '{lesson_id}')")
            continue
        title = str(
            data.get('title') or
            data.get('title ') or
            data.get('lesson_name') or
            ''
        ).strip()
        print(f"Row {row_count}: lesson_id={lesson_id}, title='{title}'")
            
        asset = clean_lesson_value(
            data.get('Asset Link') or
            data.get('Asset L') or
            ''
        )
        cocos_lesson_code = clean_lesson_value(data.get('cocos_lesson_code'))
        lido_lesson_id = get_lido_lesson_id(
            cocos_lesson_code,
            data.get('lido_lesson_id'),
        )

        if not title or title.lower() == 'nan':
            title = f"Lesson {lesson_id}"
            print(f"Row {row_count}: Using default title: {title}")
            
        if not cocos_lesson_code and not lido_lesson_id:
            print(f"Skipping row {row_count}: Missing Cocos and Lido lesson IDs")
            continue

        if cocos_lesson_code and not asset:
            print(f"Skipping row {row_count}: Missing Cocos asset link")
            continue

        valid_lessons += 1
        print(f"Processing lesson {lesson_id}: {title}")

        lesson_filename = f"{lesson_id}.json"

        lesson_manifest = create_lesson_manifest(data, lesson_id, title, asset)
        # RESPECT launches the installed Cuba Lido player. The canonical xAPI
        # activity remains the UUID URL above; the launch parameter must be
        # the Lido bundle ID, never the Cocos bundle code.
        create_tincan_xml(lesson_id, title, lido_lesson_id)

        with open(os.path.join(LESSON_DIR, lesson_filename), 'w', encoding='utf-8') as lf:
            json.dump(lesson_manifest, lf, indent=2)
        print(f"Generated lesson manifest: {lesson_filename}")

        # Get image path for OPDS
        image_filename = get_image_path(cocos_lesson_code or 'default')
        browser_launch_url = get_lido_browser_launch_url(cocos_lesson_code, lido_lesson_id)

        # For OPDS catalog, create a simplified publication entry
        publication = {
            'metadata': {
                'title': lesson_manifest['metadata']['title'],
                'author': 'Chimple',
                'identifier': f"{BASE_URL}activities/{lesson_id}",
                'language': 'en',
                'modified': lesson_manifest['metadata']['modified']
            },
            'links': [
                {
                    'rel': 'self',
                    'href': f"{BASE_URL}lessons/{lesson_id}.json",
                    'type': PUB_TYPE
                },
                {
                    'rel': 'https://id.openeel.org/rel/tincanxml',
                    'href': f"{BASE_URL}lessons/{lesson_id}/tincan.xml",
                    'type': 'application/xml'
                },
                {
                    'rel': 'https://id.openeel.org/rel/launchable-app',
                    'href': f"{BASE_URL}launchable-app.json",
                    'type': PUB_TYPE
                },
            ],
            'images': [
                {
                    'href': f"{BASE_URL}images/icons/{image_filename}",
                    'type': 'image/png',
                    'height': 128,
                    'width': 128
                }
            ]
        }

        if browser_launch_url:
            publication['links'].append({
                'rel': 'http://opds-spec.org/acquisition/open-access',
                'href': browser_launch_url,
                'type': 'text/html'
            })
        
        cocos_chapter_code = str(data.get('cocosChapterCode', '')).strip() or data.get('cocos_chapter_code', '')
        if cocos_chapter_code and cocos_chapter_code.lower() != 'nan':
             publication['metadata']['subject'] = [
                {
                    'name': title,
                    'scheme': "https://chimple.cc/curriculum",
                    'code': cocos_chapter_code
                }
            ]
        publications.append(publication)

    grade_json = {
        'metadata': {'title': f"{sheet_name}"},
        'links': [{'rel': 'self', 'href': urljoin(BASE_URL + "/grades/", grade_file), 'type': TYPE_OPDS}],
        'publications': publications
    }
    with open(os.path.join(GRADE_DIR, grade_file), 'w', encoding='utf-8') as gf:
        json.dump(grade_json, gf, indent=2)
    print(f"Generated {grade_file} with {len(publications)} valid lessons (skipped {row_count - valid_lessons} rows)")

print("\nAll files generated successfully!")
